"""Adapter for parsing Excel documents."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.cell import Cell

from automation.domain.models import Invoice
from automation.ports.document_parser import ParseResult


class ExcelInvoiceParser:
    """Parser for extracting invoice data from Excel files."""

    def can_parse(self, file_path: Path) -> bool:
        """Check whether the parser can process the file."""
        return file_path.suffix.lower() in [".xlsx", ".xls"]

    def parse_invoice(self, file_path: Path) -> ParseResult:
        """Parse an invoice from an Excel file."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Try to find invoice data across all sheets
            invoice = None
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                invoice = self._extract_invoice_from_sheet(sheet, file_path.name)
                if invoice:
                    break

            workbook.close()

            if invoice:
                return ParseResult(
                    success=True,
                    invoice=invoice,
                    metadata={"sheets_processed": len(workbook.sheetnames)},
                )
            else:
                return ParseResult(success=False, errors=["Invoice data not found in any sheet"])

        except Exception as e:
            return ParseResult(success=False, errors=[f"Excel parsing error: {str(e)}"])

    def extract_text(self, file_path: Path) -> str:
        """Extract text from an Excel file (simplified)."""
        text_lines: list[str] = []
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_values = [str(v).strip() for v in row if v is not None]
                    if row_values:
                        text_lines.append(" ".join(row_values))
        finally:
            workbook.close()
        return "\n".join(text_lines)

    def _extract_invoice_from_sheet(self, sheet, source_filename: str) -> Optional[Invoice]:
        """Extract invoice data from one Excel sheet."""

        # Search key fields across cells
        invoice_data = {
            "invoice_number": None,
            "amount": None,
            "date": None,
            "partner_id": None,
            "currency": "EUR",
        }

        # Scan first 20 rows and 10 columns
        for row in range(1, min(21, sheet.max_row + 1)):
            for col in range(1, min(11, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)

                if cell.value is None:
                    continue

                cell_str = str(cell.value).strip()

                # Invoice number
                if not invoice_data["invoice_number"]:
                    invoice_data["invoice_number"] = self._extract_invoice_number(cell_str)

                # Amount
                if not invoice_data["amount"]:
                    invoice_data["amount"] = self._extract_amount(cell)

                # Date
                if not invoice_data["date"]:
                    invoice_data["date"] = self._extract_date(cell)

                # Partner
                if not invoice_data["partner_id"]:
                    invoice_data["partner_id"] = self._extract_partner(cell_str)

        # Validate and build Invoice
        if self._is_valid_invoice_data(invoice_data):
            try:
                return Invoice(
                    partner_id=invoice_data["partner_id"] or "excel_unknown",
                    invoice_number=invoice_data["invoice_number"],
                    invoice_date=invoice_data["date"] or datetime.now().date(),
                    amount=invoice_data["amount"],
                    currency=invoice_data["currency"],
                    source_message_id=source_filename,
                )
            except (ValueError, TypeError):
                return None

        return None

    def _extract_invoice_number(self, text: str) -> Optional[str]:
        """Extract invoice number from cell text."""
        import re

        # Invoice number patterns
        patterns = [
            r"(?:Invoice|Bill|Inv|№)[\s#:]*([A-Z0-9/-]+)",
        ]

        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                invoice_num = match.group(1).strip()
                if len(invoice_num) > 2:  # Minimal number length
                    return invoice_num

        return None

    def _extract_amount(self, cell: Cell) -> Optional[Decimal]:
        """Extract amount from an Excel cell."""
        value = cell.value

        # Numeric cell
        if isinstance(value, (int, float)):
            if value > 0:
                return Decimal(str(value))

        # String cell with embedded number
        elif isinstance(value, str):
            import re

            # Examples: 123.45, 1,234.56, 1 234,56
            amount_match = re.search(r"([0-9]{1,3}(?:[,\s][0-9]{3})*(?:\.[0-9]{2})?)", value)
            if amount_match:
                amount_str = amount_match.group(1).replace(",", "").replace(" ", "")
                try:
                    amount = Decimal(amount_str)
                    if amount > 0:
                        return amount
                except (ValueError, TypeError):
                    pass

        return None

    def _extract_date(self, cell: Cell) -> Optional[date]:
        """Extract date from an Excel cell."""
        value = cell.value

        # Excel datetime
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, date):
            return value

        # Date represented as text
        elif isinstance(value, str):
            return self._parse_date_string(value)

        return None

    def _parse_date_string(self, date_str: str) -> Optional[date]:
        """Parse date from a string."""
        import re

        # Supported formats: DD.MM.YYYY, DD/MM/YYYY, YYYY-MM-DD
        date_patterns = [
            (r"(\d{1,2})[./](\d{1,2})[./](\d{4})", "%d/%m/%Y"),
            (r"(\d{4})-(\d{1,2})-(\d{1,2})", "%Y-%m-%d"),
        ]

        for pattern, fmt in date_patterns:
            match = re.search(pattern, date_str)
            if match:
                try:
                    if fmt == "%d/%m/%Y":
                        day, month, year = match.groups()
                        return datetime(int(year), int(month), int(day)).date()
                    elif fmt == "%Y-%m-%d":
                        year, month, day = match.groups()
                        return datetime(int(year), int(month), int(day)).date()
                except (ValueError, TypeError):
                    continue

        return None

    def _extract_partner(self, text: str) -> Optional[str]:
        """Extract partner identifier from text."""
        # Simple heuristic for partner detection
        keywords = ["supplier", "vendor", "seller", "from"]

        text_lower = text.lower()
        for keyword in keywords:
            if keyword in text_lower:
                # Extract word after keyword
                import re

                pattern = rf"{keyword}[\s:]*([a-zA-Z]+)"
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    return match.group(1).lower()

        return None

    def _is_valid_invoice_data(self, data: dict) -> bool:
        """Check minimum validity for extracted invoice data."""
        return (
            data.get("invoice_number") is not None
            and data.get("amount") is not None
            and data["amount"] > 0
        )
